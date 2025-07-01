import pandas as pd
import random
import json
import os
import re
import shutil
import pickle
import atexit
import traceback
from datetime import datetime, time, timedelta
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from collections import defaultdict

class ScheduleGenerator:
    def __init__(self):
        self.lecturers = []
        self.subjects = []
        self.classes = []
        self.fixed_schedules = []
        self.generated_schedules = []
        self.available_rooms = []
        self.break_times = [
            {"start": time(12, 0), "end": time(13, 0)},
            {"start": time(18, 0), "end": time(19, 0)}
        ]
        self.additional_break_times = [
            {"start": time(12, 11), "end": time(12, 59)},
            {"start": time(19, 10), "end": time(19, 59)}
        ]
        self.department_preferences = {
            "TI": [3, 4],
            "SI": [3, 4],
            "DKV": [5],
            "default": [3, 4, 5]
        }
        self.room_capacities = {}
        self.time_slots = self.generate_time_slots()
        self.excel_path = None
        self.lecturer_breaks = defaultdict(list)
        self.lecturer_preferences = defaultdict(dict)
        self.cache_file = "schedule_cache.pkl"
        self.days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
        self.max_attempts = 200
        self.ui_state_file = "ui_state.json"
        self.online_ratio = 0.2  # Rasio 20% online, 80% offline

    def generate_time_slots(self):
        slots = []
        start_hour = 8
        end_hour = 21
        
        # Generate offline and online slots for all hours
        for sks in [1, 2, 3, 4]:
            for hour in range(start_hour, end_hour):
                total_minutes = hour * 60 + (sks * 50)
                end_hour_val = total_minutes // 60
                end_minute_val = total_minutes % 60
                
                if end_hour_val >= 21:
                    continue
                    
                # Offline slot
                start_str = f"{hour:02d}:00"
                end_str = f"{end_hour_val:02d}:{end_minute_val:02d}"
                
                if (start_str, end_str) not in slots:
                    slots.append((start_str, end_str))
                
                # Online slot
                online_start_str = f"{hour:02d}:00 (online)"
                online_end_str = f"{end_hour_val:02d}:{end_minute_val:02d} (online)"
                
                if (online_start_str, online_end_str) not in slots:
                    slots.append((online_start_str, online_end_str))
        
        # Add special online slots
        online_slots = [
            ("15:30 (online)", "17:10 (online)"),
            ("17:40 (online)", "19:20 (online)"),
            ("19:30 (online)", "21:10 (online)")
        ]
        
        return sorted(slots + online_slots, key=lambda x: (
            datetime.strptime(x[0].replace(' (online)', ''), "%H:%M"),
            datetime.strptime(x[1].replace(' (online)', ''), "%H:%M")
        ))

    def is_valid_for_sks(self, time_slot, sks):
        try:
            start_str = time_slot[0].replace(' (online)', '')
            end_str = time_slot[1].replace(' (online)', '')
            
            start_time = datetime.strptime(start_str, "%H:%M")
            end_time = datetime.strptime(end_str, "%H:%M")
            
            duration = (end_time - start_time).total_seconds() / 60
            return abs(duration - (sks * 50)) < 1.0
        except:
            return False

    def save_cache(self):
        try:
            with open(self.cache_file, 'wb') as f:
                data = {
                    'lecturers': self.lecturers,
                    'subjects': self.subjects,
                    'classes': self.classes,
                    'fixed_schedules': self.fixed_schedules,
                    'generated_schedules': self.generated_schedules,
                    'lecturer_breaks': dict(self.lecturer_breaks),
                    'lecturer_preferences': dict(self.lecturer_preferences),
                    'excel_path': self.excel_path,
                    'available_rooms': self.available_rooms,
                    'room_capacities': self.room_capacities
                }
                pickle.dump(data, f)
        except Exception as e:
            print(f"Error saving cache: {e}")

    def load_cache(self):
        try:
            if os.path.exists(self.cache_file):
                with open(self.cache_file, 'rb') as f:
                    data = pickle.load(f)
                    self.lecturers = data.get('lecturers', [])
                    self.subjects = data.get('subjects', [])
                    self.classes = data.get('classes', [])
                    self.fixed_schedules = data.get('fixed_schedules', [])
                    self.generated_schedules = data.get('generated_schedules', [])
                    self.lecturer_breaks = defaultdict(list, data.get('lecturer_breaks', {}))
                    self.lecturer_preferences = defaultdict(dict, data.get('lecturer_preferences', {}))
                    self.excel_path = data.get('excel_path')
                    self.available_rooms = data.get('available_rooms', [])
                    self.room_capacities = data.get('room_capacities', {})
                return True
        except:
            pass
        return False

    def save_ui_state(self, state):
        try:
            with open(self.ui_state_file, 'w') as f:
                json.dump(state, f)
        except Exception as e:
            print(f"Error saving UI state: {e}")

    def load_ui_state(self):
        try:
            if os.path.exists(self.ui_state_file):
                with open(self.ui_state_file, 'r') as f:
                    return json.load(f)
        except:
            pass
        return {}

    def add_lecturer_preference(self, lecturer, 
                               available_days=None, 
                               preferred_times_offline=None, 
                               preferred_times_online=None, 
                               online_days=None, 
                               use_additional_breaks=False):
        available_days = available_days or []
        preferred_times_offline = preferred_times_offline or []
        preferred_times_online = preferred_times_online or []
        online_days = online_days or []
        
        valid_prefs = {
            'available_days': [],
            'online_days': [],
            'preferred_times_offline': [],
            'preferred_times_online': [],
            'use_additional_breaks': use_additional_breaks
        }
        
        for day in available_days:
            if day in self.days:
                valid_prefs['available_days'].append(day)
        
        for day in online_days:
            if day in self.days and day not in valid_prefs['available_days']:
                valid_prefs['online_days'].append(day)
        
        for start, end in preferred_times_offline:
            if self.is_valid_time_range(start, end):
                valid_prefs['preferred_times_offline'].append((start, end))
        
        for start, end in preferred_times_online:
            if self.is_valid_time_range(start, end):
                valid_prefs['preferred_times_online'].append((start, end))
        
        self.lecturer_preferences[lecturer] = valid_prefs

    def parse_time(self, time_str):
        try:
            time_str = str(time_str).strip()
            is_online = "(online)" in time_str.lower()
            time_part = re.sub(r'\(.*\)', '', time_str).strip().replace('.', ':')
            
            if ':' in time_part:
                hours, minutes = time_part.split(':')[:2]
                time_part = f"{hours}:{minutes[:2]}"
            
            return datetime.strptime(time_part, "%H:%M").time(), is_online
        except:
            return None, False

    def is_valid_time_range(self, start, end):
        start_time, _ = self.parse_time(start)
        end_time, _ = self.parse_time(end)
        return start_time and end_time and start_time < end_time

    def is_break_time(self, start_str, end_str):
        start_time, _ = self.parse_time(start_str)
        end_time, _ = self.parse_time(end_str)
        if not start_time or not end_time:
            return False
            
        for bt in self.break_times:
            if start_time < bt['end'] and end_time > bt['start']:
                return True
        return False

    def load_data(self, excel_path):
        try:
            self.excel_path = excel_path
            df = pd.read_excel(excel_path, sheet_name='Mapping mata kuliah', skiprows=2)
            df = df.dropna(subset=['Nama Dosen', 'Mata Kuliah'])

            # Konversi kolom SKS dan Semester ke integer
            df['SKS'] = df['SKS'].fillna(0).astype(int)
            df['Semester'] = df['Semester'].fillna(0).astype(int)
        
            self.lecturers = df['Nama Dosen'].unique().tolist()
            self.subjects = df['Mata Kuliah'].unique().tolist()
            self.classes = df['Kelas'].unique().tolist()
            self.fixed_schedules = []
        
            for idx, row in df.iterrows():
                self.fixed_schedules.append({
                    'source': 'excel',
                    'excel_index': idx,
                    'dosen': row['Nama Dosen'],
                    'mata_kuliah': row['Mata Kuliah'],
                    'kelas': row['Kelas'],
                    'hari': "",
                    'jam': "",
                    'semester': row['Semester'],  # Sudah integer
                    'sks': row['SKS'],            # Sudah integer
                    'ruangan': "",
                    'jumlah_mahasiswa': row.get('Jumlah Mahasiswa', 0),
                    'is_fixed': False  # Default tidak tetap
                })
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Gagal memuat data: {str(e)}")
            return False

    def load_rooms(self, json_path):
        try:
            with open(json_path, 'r') as f:
                rooms = json.load(f)
                self.available_rooms = [room for room in rooms if 'online' not in room['nama'].lower()]
                self.room_capacities = {room['nama']: room.get('kapasitas', 30) for room in self.available_rooms}
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Gagal memuat data ruangan: {str(e)}")
            return False

    def is_time_overlap(self, start1, end1, start2, end2):
        return not (end1 <= start2 or start1 >= end2)

    def is_conflict(self, schedule, check_room_capacity=True):
        if not schedule['jam']:
            return False
            
        jam_parts = schedule['jam'].split(' - ')
        if len(jam_parts) != 2:
            return True
            
        start, end = jam_parts
        start_time, is_online = self.parse_time(start)
        end_time, _ = self.parse_time(end)
        
        if not start_time or not end_time or not self.is_valid_time_range(start, end):
            return True
            
        all_schedules = self.fixed_schedules + self.generated_schedules
        
        # 1. Check lecturer availability
        for sched in all_schedules:
            if (sched != schedule and sched['dosen'] == schedule['dosen'] 
                and sched['hari'] == schedule['hari'] and sched['jam']):
                s_start, s_end = sched['jam'].split(' - ')
                s_start_time, _ = self.parse_time(s_start)
                s_end_time, _ = self.parse_time(s_end)
                
                if s_start_time and s_end_time and self.is_time_overlap(start_time, end_time, s_start_time, s_end_time):
                    return True
        
        # Check lecturer preferences
        lecturer_pref = self.lecturer_preferences.get(schedule['dosen'], {})
        
        # Check available days (hari yang tersedia)
        available_days = lecturer_pref.get('available_days', [])
        if available_days and schedule['hari'] not in available_days:
            return True
            
        # Check online days
        if schedule['hari'] in lecturer_pref.get('online_days', []) and schedule.get('ruangan') != 'Online':
            return True
            
        # Check preferred times (online/offline specific)
        if lecturer_pref.get('preferred_times_offline') or lecturer_pref.get('preferred_times_online'):
            time_ok = False
            is_online_class = schedule.get('ruangan') == 'Online'
            
            if is_online_class:
                preferred_times = lecturer_pref.get('preferred_times_online', [])
            else:
                preferred_times = lecturer_pref.get('preferred_times_offline', [])
            
            for pref_start, pref_end in preferred_times:
                pref_start_time, _ = self.parse_time(pref_start)
                pref_end_time, _ = self.parse_time(pref_end)
                
                if pref_start_time and pref_end_time and start_time >= pref_start_time and end_time <= pref_end_time:
                    time_ok = True
                    break
            
            if preferred_times and not time_ok:
                return True
                    
        # 2. Check room availability and capacity (only for offline classes)
        if (check_room_capacity and schedule.get('ruangan') 
            and schedule['ruangan'] != 'Online' and schedule.get('jam')):
            for sched in all_schedules:
                if (sched != schedule and sched.get('ruangan') == schedule['ruangan'] 
                    and sched['hari'] == schedule['hari'] and sched['jam']):
                    s_start, s_end = sched['jam'].split(' - ')
                    s_start_time, _ = self.parse_time(s_start)
                    s_end_time, _ = self.parse_time(s_end)
                    
                    if s_start_time and s_end_time and self.is_time_overlap(start_time, end_time, s_start_time, s_end_time):
                        return True
            
            room_capacity = self.room_capacities.get(schedule['ruangan'], 0)
            if schedule.get('jumlah_mahasiswa', 0) > room_capacity:
                return True
        
        # 3. Check class availability (for both online and offline)
        for sched in all_schedules:
            if (sched != schedule and sched['kelas'] == schedule['kelas'] 
                and sched['hari'] == schedule['hari'] and sched['jam']):
                s_start, s_end = sched['jam'].split(' - ')
                s_start_time, _ = self.parse_time(s_start)
                s_end_time, _ = self.parse_time(s_end)
                
                if s_start_time and s_end_time and self.is_time_overlap(start_time, end_time, s_start_time, s_end_time):
                    return True
        
        # 4. Check break times (only for offline classes)
        if schedule.get('ruangan') != 'Online' and schedule.get('jam'):
            break_times_to_check = self.break_times.copy()
            
            if lecturer_pref.get('use_additional_breaks', False):
                break_times_to_check.extend(self.additional_break_times)
            
            for bt in break_times_to_check:
                if start_time < bt['end'] and end_time > bt['start']:
                    return True
            
        # 5. Check lecturer break times (for both online and offline)
        lecturer_breaks = self.lecturer_breaks.get(schedule['dosen'], [])
        for break_time in lecturer_breaks:
            break_start, break_end = break_time.split(' - ')
            break_start_time, _ = self.parse_time(break_start)
            break_end_time, _ = self.parse_time(break_end)
            
            if break_start_time and break_end_time and self.is_time_overlap(start_time, end_time, break_start_time, break_end_time):
                return True
                
        return False

    def get_conflict_reasons(self, schedule):
        reasons = []
        if not schedule.get('jam'):
            return ["Jadwal belum diisi waktu"]
        
        jam_parts = schedule['jam'].split(' - ')
        if len(jam_parts) != 2:
            return ["Format waktu tidak valid"]
        
        start, end = jam_parts
        start_time, is_online = self.parse_time(start)
        end_time, _ = self.parse_time(end)
        
        if not start_time or not end_time or not self.is_valid_time_range(start, end):
            reasons.append("Format waktu tidak valid")
            return reasons
        
        all_schedules = self.fixed_schedules + self.generated_schedules
        
        # Check lecturer conflict
        for sched in all_schedules:
            if (sched != schedule and sched['dosen'] == schedule['dosen'] 
                and sched['hari'] == schedule['hari'] and sched['jam']):
                s_start, s_end = sched['jam'].split(' - ')
                s_start_time, _ = self.parse_time(s_start)
                s_end_time, _ = self.parse_time(s_end)
                
                if s_start_time and s_end_time and self.is_time_overlap(start_time, end_time, s_start_time, s_end_time):
                    reasons.append(f"Konflik dengan dosen di jadwal {sched['mata_kuliah']} (kelas {sched['kelas']})")
        
        # Check room conflict and capacity (only for offline)
        if schedule.get('ruangan') and schedule['ruangan'] != 'Online':
            room_capacity = self.room_capacities.get(schedule['ruangan'], 0)
            if schedule.get('jumlah_mahasiswa', 0) > room_capacity:
                reasons.append(f"Kapasitas ruangan {schedule['ruangan']} ({room_capacity}) terlampaui")
            
            for sched in all_schedules:
                if (sched != schedule and sched.get('ruangan') == schedule['ruangan'] 
                    and sched['hari'] == schedule['hari'] and sched['jam']):
                    s_start, s_end = sched['jam'].split(' - ')
                    s_start_time, _ = self.parse_time(s_start)
                    s_end_time, _ = self.parse_time(s_end)
                    
                    if s_start_time and s_end_time and self.is_time_overlap(start_time, end_time, s_start_time, s_end_time):
                        reasons.append(f"Konflik ruangan dengan jadwal {sched['mata_kuliah']} (kelas {sched['kelas']})")
        
        # Check class conflict
        for sched in all_schedules:
            if (sched != schedule and sched['kelas'] == schedule['kelas'] 
                and sched['hari'] == schedule['hari'] and sched['jam']):
                s_start, s_end = sched['jam'].split(' - ')
                s_start_time, _ = self.parse_time(s_start)
                s_end_time, _ = self.parse_time(s_end)
                
                if s_start_time and s_end_time and self.is_time_overlap(start_time, end_time, s_start_time, s_end_time):
                    reasons.append(f"Konflik kelas dengan jadwal {sched['mata_kuliah']}")
        
        # Check break times (only for offline)
        if schedule.get('ruangan') != 'Online':
            lecturer_pref = self.lecturer_preferences.get(schedule['dosen'], {})
            break_times_to_check = self.break_times.copy()
            if lecturer_pref.get('use_additional_breaks', False):
                break_times_to_check.extend(self.additional_break_times)
            
            for bt in break_times_to_check:
                if start_time < bt['end'] and end_time > bt['start']:
                    reasons.append(f"Tumpang tindih dengan waktu istirahat ({bt['start'].strftime('%H:%M')}-{bt['end'].strftime('%H:%M')})")
        
        # Check lecturer breaks
        lecturer_breaks = self.lecturer_breaks.get(schedule['dosen'], [])
        for break_time in lecturer_breaks:
            break_start, break_end = break_time.split(' - ')
            break_start_time, _ = self.parse_time(break_start)
            break_end_time, _ = self.parse_time(break_end)
            
            if break_start_time and break_end_time and self.is_time_overlap(start_time, end_time, break_start_time, break_end_time):
                reasons.append(f"Tumpang tindih dengan waktu istirahat dosen ({break_start}-{break_end})")
        
        # Check lecturer preferences
        available_days = lecturer_pref.get('available_days', [])
        if available_days and schedule['hari'] not in available_days:
            reasons.append(f"Hari {schedule['hari']} tidak tersedia untuk dosen ini")
            
        if schedule['hari'] in lecturer_pref.get('online_days', []) and schedule.get('ruangan') != 'Online':
            reasons.append(f"Hari {schedule['hari']} harus online tetapi jadwal offline")
            
        if lecturer_pref.get('preferred_times_offline') or lecturer_pref.get('preferred_times_online'):
            time_ok = False
            is_online_class = schedule.get('ruangan') == 'Online'
            
            if is_online_class:
                preferred_times = lecturer_pref.get('preferred_times_online', [])
            else:
                preferred_times = lecturer_pref.get('preferred_times_offline', [])
            
            for pref_start, pref_end in preferred_times:
                pref_start_time, _ = self.parse_time(pref_start)
                pref_end_time, _ = self.parse_time(pref_end)
                
                if pref_start_time and pref_end_time and start_time >= pref_start_time and end_time <= pref_end_time:
                    time_ok = True
                    break
            
            if preferred_times and not time_ok:
                reasons.append("Waktu tidak sesuai preferensi dosen")
        
        return reasons

    def get_available_room(self, department, day, start_time_str, end_time_str, student_count=0):
        try:
            start_time, is_online = self.parse_time(start_time_str)
            end_time, _ = self.parse_time(end_time_str)
            
            if is_online or not start_time or not end_time:
                return 'Online' if is_online else None
                
            lecturer_pref = self.lecturer_preferences.get(department, {})
            if day in lecturer_pref.get('online_days', []):
                return 'Online'
                
            preferred_floors = self.department_preferences.get(department, self.department_preferences['default'])
            rooms = self.available_rooms.copy()
            random.shuffle(rooms)
            
            valid_rooms = [room for room in rooms 
                          if room.get('kapasitas', 30) >= student_count]
            
            for room in valid_rooms:
                if room.get('lantai') in preferred_floors:
                    available = True
                    for sched in self.fixed_schedules + self.generated_schedules:
                        if (sched.get('ruangan') == room['nama'] 
                            and sched['hari'] == day 
                            and sched['jam']):
                            s_start, s_end = sched['jam'].split(' - ')
                            s_start_time, _ = self.parse_time(s_start)
                            s_end_time, _ = self.parse_time(s_end)
                            
                            if (s_start_time and s_end_time 
                                and self.is_time_overlap(start_time, end_time, s_start_time, s_end_time)):
                                available = False
                                break
                    if available:
                        return room['nama']
                    
            for room in valid_rooms:
                available = True
                for sched in self.fixed_schedules + self.generated_schedules:
                    if (sched.get('ruangan') == room['nama'] 
                        and sched['hari'] == day 
                        and sched['jam']):
                        s_start, s_end = sched['jam'].split(' - ')
                        s_start_time, _ = self.parse_time(s_start)
                        s_end_time, _ = self.parse_time(s_end)
                        
                        if (s_start_time and s_end_time 
                            and self.is_time_overlap(start_time, end_time, s_start_time, s_end_time)):
                            available = False
                            break
                if available:
                    return room['nama']
                    
            return None
        except Exception as e:
            print(f"Error in get_available_room: {e}")
            return None

    def fill_empty_rooms_randomly(self):
        try:
            all_schedules = self.fixed_schedules + self.generated_schedules
            schedules_without_room = [
                s for s in all_schedules 
                if not s.get('ruangan') or str(s.get('ruangan')).strip() == ''
            ]
            
            schedules_without_room.sort(key=lambda x: x.get('jumlah_mahasiswa', 0), reverse=True)
            
            for sched in schedules_without_room:
                is_online = "(online)" in str(sched.get('jam', '')).lower()
                lecturer_pref = self.lecturer_preferences.get(sched['dosen'], {})
                
                if is_online or sched['hari'] in lecturer_pref.get('online_days', []):
                    sched['ruangan'] = 'Online'
                    continue
                    
                if not sched.get('jam'):
                    continue
                    
                department = sched['kelas'][:2] if isinstance(sched['kelas'], str) and len(sched['kelas']) >= 2 else 'default'
                jam_parts = sched['jam'].split(' - ')
                if len(jam_parts) != 2:
                    continue
                    
                start, end = jam_parts
                student_count = sched.get('jumlah_mahasiswa', 0)
                
                room = self.get_available_room(
                    department, 
                    sched['hari'], 
                    start, 
                    end,
                    student_count
                )
                if room:
                    sched['ruangan'] = room
                else:
                    for room in self.available_rooms:
                        if room.get('kapasitas', 30) < student_count:
                            continue
                            
                        available = True
                        for existing in all_schedules:
                            if (existing.get('ruangan') == room['nama'] 
                                and existing['hari'] == sched['hari'] 
                                and existing.get('jam')):
                                try:
                                    e_jam_parts = existing['jam'].split(' - ')
                                    if len(e_jam_parts) != 2:
                                        continue
                                    e_start, e_end = e_jam_parts
                                    e_start_time, _ = self.parse_time(e_start)
                                    e_end_time, _ = self.parse_time(e_end)
                                    current_start_time, _ = self.parse_time(start)
                                    current_end_time, _ = self.parse_time(end)
                                    
                                    if (e_start_time and e_end_time and current_start_time and current_end_time
                                        and self.is_time_overlap(current_start_time, current_end_time, e_start_time, e_end_time)):
                                        available = False
                                        break
                                except:
                                    continue
                        if available:
                            sched['ruangan'] = room['nama']
                            break
            return True
        except Exception as e:
            print(f"Error in fill_empty_rooms_randomly: {e}")
            return False

    def save_to_excel(self, schedules, template_path, output_folder):
        try:
            output_path = os.path.join(output_folder, f"Jadwal_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            wb = load_workbook(template_path)
            sheet = wb.active
            
            headers = ["Hari", "Mata Kuliah", "Kelas", "Ruangan", "Jam", "SKS", "Semester", "Dosen", "Jumlah Mahasiswa"]
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=3, column=col, value=header)
            
            for row_idx, sched in enumerate(schedules, start=4):
                sheet.cell(row=row_idx, column=1, value=sched['hari'])
                sheet.cell(row=row_idx, column=2, value=sched['mata_kuliah'])
                sheet.cell(row=row_idx, column=3, value=sched['kelas'])
                sheet.cell(row=row_idx, column=4, value=sched.get('ruangan', ''))
                sheet.cell(row=row_idx, column=5, value=sched['jam'])
                sheet.cell(row=row_idx, column=6, value=sched['sks'])
                sheet.cell(row=row_idx, column=7, value=sched['semester'])
                sheet.cell(row=row_idx, column=8, value=sched['dosen'])
                sheet.cell(row=row_idx, column=9, value=sched.get('jumlah_mahasiswa', ''))
                
            wb.save(output_path)
            return output_path
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menyimpan: {str(e)}")
            return None

    def find_all_conflicts(self):
        conflicts = {
            'lecturer': [],
            'room': [],
            'capacity': [],
            'class': [],
            'break_time': [],
            'online_day': [],
            'preference': []
        }
        
        all_schedules = self.fixed_schedules + self.generated_schedules
        
        for i, sched in enumerate(all_schedules):
            if not sched.get('jam'):
                continue
                
            for j, other in enumerate(all_schedules[i+1:], i+1):
                if (sched['dosen'] == other['dosen'] 
                    and sched['hari'] == other['hari'] 
                    and other.get('jam')):
                    try:
                        s_start, s_end = sched['jam'].split(' - ')
                        o_start, o_end = other['jam'].split(' - ')
                        
                        s_start_time, _ = self.parse_time(s_start)
                        s_end_time, _ = self.parse_time(s_end)
                        o_start_time, _ = self.parse_time(o_start)
                        o_end_time, _ = self.parse_time(o_end)
                        
                        if (s_start_time and s_end_time and o_start_time and o_end_time
                            and self.is_time_overlap(s_start_time, s_end_time, o_start_time, o_end_time)):
                            conflicts['lecturer'].append({
                                'conflict_type': 'Dosen ganda',
                                'dosen': sched['dosen'],
                                'hari': sched['hari'],
                                'waktu': f"{max(s_start_time, o_start_time)}-{min(s_end_time, o_end_time)}",
                                'schedule1': sched,
                                'schedule2': other
                            })
                    except:
                        continue
            
            if (sched.get('ruangan') and sched['ruangan'] != 'Online' 
                and sched.get('jam')):
                for j, other in enumerate(all_schedules[i+1:], i+1):
                    if (other.get('ruangan') == sched['ruangan'] 
                        and other['hari'] == sched['hari'] 
                        and other.get('jam')):
                        try:
                            s_start, s_end = sched['jam'].split(' - ')
                            o_start, o_end = other['jam'].split(' - ')
                            
                            s_start_time, _ = self.parse_time(s_start)
                            s_end_time, _ = self.parse_time(s_end)
                            o_start_time, _ = self.parse_time(o_start)
                            o_end_time, _ = self.parse_time(o_end)
                            
                            if (s_start_time and s_end_time and o_start_time and o_end_time
                                and self.is_time_overlap(s_start_time, s_end_time, o_start_time, o_end_time)):
                                conflicts['room'].append({
                                    'conflict_type': 'Ruangan ganda',
                                    'ruangan': sched['ruangan'],
                                    'hari': sched['hari'],
                                    'waktu': f"{max(s_start_time, o_start_time)}-{min(s_end_time, o_end_time)}",
                                    'schedule1': sched,
                                    'schedule2': other
                                })
                        except:
                            continue
                
                room_capacity = self.room_capacities.get(sched['ruangan'], 0)
                if sched.get('jumlah_mahasiswa', 0) > room_capacity:
                    conflicts['capacity'].append({
                        'conflict_type': 'Kapasitas ruangan terlampaui',
                        'ruangan': sched['ruangan'],
                        'kapasitas': room_capacity,
                        'mahasiswa': sched.get('jumlah_mahasiswa', 0),
                        'schedule': sched
                    })
            
            for j, other in enumerate(all_schedules[i+1:], i+1):
                if (sched['kelas'] == other['kelas'] 
                    and sched['hari'] == other['hari'] 
                    and other.get('jam')):
                    try:
                        s_start, s_end = sched['jam'].split(' - ')
                        o_start, o_end = other['jam'].split(' - ')
                        
                        s_start_time, _ = self.parse_time(s_start)
                        s_end_time, _ = self.parse_time(s_end)
                        o_start_time, _ = self.parse_time(o_start)
                        o_end_time, _ = self.parse_time(o_end)
                        
                        if (s_start_time and s_end_time and o_start_time and o_end_time
                            and self.is_time_overlap(s_start_time, s_end_time, o_start_time, o_end_time)):
                            conflicts['class'].append({
                                'conflict_type': 'Kelas ganda',
                                'kelas': sched['kelas'],
                                'hari': sched['hari'],
                                'waktu': f"{max(s_start_time, o_start_time)}-{min(s_end_time, o_end_time)}",
                                'schedule1': sched,
                                'schedule2': other
                            })
                    except:
                        continue
            
            if sched.get('ruangan') != 'Online' and sched.get('jam'):
                try:
                    jam_parts = sched['jam'].split(' - ')
                    if len(jam_parts) == 2:
                        start, end = jam_parts
                        if self.is_break_time(start, end):
                            conflicts['break_time'].append({
                                'conflict_type': 'Waktu istirahat',
                                'dosen': sched['dosen'],
                                'hari': sched['hari'],
                                'waktu': sched['jam'],
                                'schedule': sched
                            })
                except:
                    pass
                
            lecturer_pref = self.lecturer_preferences.get(sched['dosen'], {})
            online_days = lecturer_pref.get('online_days', [])
            if sched['hari'] in online_days and sched.get('ruangan') != 'Online':
                conflicts['online_day'].append({
                    'conflict_type': 'Hari online tidak menggunakan ruang online',
                    'dosen': sched['dosen'],
                    'hari': sched['hari'],
                    'waktu': sched['jam'],
                    'schedule': sched,
                    'ruangan': sched.get('ruangan', '')
                })
            
            available_days = lecturer_pref.get('available_days', [])
            if available_days and sched['hari'] not in available_days:
                conflicts['preference'].append({
                    'conflict_type': 'Hari tidak tersedia',
                    'dosen': sched['dosen'],
                    'hari': sched['hari'],
                    'waktu': sched['jam'],
                    'schedule': sched
                })
                
            if lecturer_pref.get('preferred_times_offline') or lecturer_pref.get('preferred_times_online'):
                time_ok = False
                start_time, _ = self.parse_time(sched['jam'].split(' - ')[0])
                end_time, _ = self.parse_time(sched['jam'].split(' - ')[1])
                
                is_online = sched.get('ruangan') == 'Online'
                preferred_times = lecturer_pref.get('preferred_times_online' if is_online else 'preferred_times_offline', [])
                
                for pref_start, pref_end in preferred_times:
                    pref_start_time, _ = self.parse_time(pref_start)
                    pref_end_time, _ = self.parse_time(pref_end)
                    
                    if pref_start_time and pref_end_time and start_time >= pref_start_time and end_time <= pref_end_time:
                        time_ok = True
                        break
                
                if preferred_times and not time_ok:
                    conflicts['preference'].append({
                        'conflict_type': 'Waktu tidak diinginkan',
                        'dosen': sched['dosen'],
                        'hari': sched['hari'],
                        'waktu': sched['jam'],
                        'schedule': sched
                    })
        
        return conflicts
    
    def suggest_conflict_resolutions(self, conflict):
        suggestions = []
        
        if conflict['conflict_type'] == 'Dosen ganda':
            for day in self.days:
                if day == conflict['hari']:
                    continue
                    
                new_schedule = conflict['schedule1'].copy()
                new_schedule['hari'] = day
                
                if not self.is_conflict(new_schedule):
                    if self.edit_schedule(conflict['schedule1'], new_schedule):
                        resolved += 1
                        break
                        
        for conflict in conflict['online_day']:
            new_schedule = conflict['schedule'].copy()
            new_schedule['ruangan'] = 'Online'
            if not self.is_conflict(new_schedule):
                if self.edit_schedule(conflict['schedule'], new_schedule):
                    resolved += 1
                        
        return resolved

    def add_manual_schedule(self, schedule):
        schedule['source'] = 'manual'
        schedule['is_fixed'] = schedule.get('is_fixed', False)  # Tambahkan atribut is_fixed
        self.fixed_schedules.append(schedule)
        
        if schedule['dosen'] not in self.lecturers:
            self.lecturers.append(schedule['dosen'])
        if schedule['mata_kuliah'] not in self.subjects:
            self.subjects.append(schedule['mata_kuliah'])
        if schedule['kelas'] not in self.classes:
            self.classes.append(schedule['kelas'])
            
        return True

    def remove_schedule(self, schedule):
        if schedule in self.fixed_schedules:
            self.fixed_schedules.remove(schedule)
            return True
        elif schedule in self.generated_schedules:
            self.generated_schedules.remove(schedule)
            return True
        return False

    def edit_schedule(self, old_schedule, new_schedule):
        if self.remove_schedule(old_schedule):
            # Jadwal dari Excel diubah menjadi manual setelah diedit
            new_schedule['source'] = 'manual'
            if old_schedule.get('source') == 'excel':
                new_schedule['excel_index'] = old_schedule['excel_index']
            
            self.fixed_schedules.append(new_schedule)
            return True
        return False

    def auto_resolve_conflicts(self):
        resolved = 0
        conflicts = self.find_all_conflicts()
        
        for conflict in conflicts['lecturer']:
            for day in self.days:
                if day == conflict['hari']:
                    continue
                    
                new_schedule = conflict['schedule1'].copy()
                new_schedule['hari'] = day
                
                if not self.is_conflict(new_schedule):
                    if self.edit_schedule(conflict['schedule1'], new_schedule):
                        resolved += 1
                        break
                        
        for conflict in conflicts['online_day']:
            new_schedule = conflict['schedule'].copy()
            new_schedule['ruangan'] = 'Online'
            if not self.is_conflict(new_schedule):
                if self.edit_schedule(conflict['schedule'], new_schedule):
                    resolved += 1
                        
        return resolved

    def add_lecturer_break(self, lecturer, day, start_time, end_time):
        key = f"{lecturer}|{day}"
        self.lecturer_breaks[key].append(f"{start_time} - {end_time}")

    def randomize_schedule(self, reshuffle_existing=False):
        # Reset only schedules that have been scheduled when reshuffling
        if reshuffle_existing:
            for s in self.fixed_schedules + self.generated_schedules:
                # Skip reset jika jadwal sudah di-flag sebagai fixed
                if s.get('is_fixed', False):
                    continue
                    
                if s.get('source') == 'excel' or s.get('source') == 'manual':
                    s['hari'] = ""
                    s['jam'] = ""
                    s['ruangan'] = ""
        
        # Ambil semua jadwal yang belum terjadwal (baik excel maupun manual) dan bukan fixed
        unscheduled = [s for s in self.fixed_schedules + self.generated_schedules 
                      if (not s.get('hari') or not s.get('jam')) and not s.get('is_fixed', False)]
    
        if not unscheduled:
            return 0, 0, []  # Return empty list for failures
        
        success_count = 0
        failure_count = 0
        failed_schedules = []  # List of dictionaries with schedule and reasons
        
        unscheduled.sort(key=lambda x: x['sks'], reverse=True)
        
        for schedule in unscheduled:
            assigned = False
            valid_days = self.days.copy()
            
            # Apply lecturer preferences for available days
            lecturer_pref = self.lecturer_preferences.get(schedule['dosen'], {})
            available_days = lecturer_pref.get('available_days', [])
            if available_days:
                valid_days = [day for day in available_days]
            
            if not valid_days:
                failure_count += 1
                conflict_reasons = ["Tidak ada hari yang tersedia (dari preferensi dosen)"]
                failed_schedules.append({
                    'schedule': schedule,
                    'reasons': conflict_reasons
                })
                continue
                
            for attempt in range(self.max_attempts):
                day = random.choice(valid_days)
                
                # Determine if this should be online class based on ratio
                is_online_class = False
                lecturer_pref = self.lecturer_preferences.get(schedule['dosen'], {})
                
                # Prioritize lecturer preferences
                if day in lecturer_pref.get('online_days', []):
                    is_online_class = True
                else:
                    # Apply 20% online ratio only if not specified by lecturer
                    is_online_class = random.random() < self.online_ratio
                
                valid_slots = []
                for slot in self.time_slots:
                    if not self.is_valid_for_sks(slot, schedule['sks']):
                        continue
                    
                    slot_is_online = "(online)" in slot[0].lower() or "(online)" in slot[1].lower()
                    
                    # Filter slots based on online/offline requirement
                    if is_online_class and slot_is_online:
                        valid_slots.append(slot)
                    elif not is_online_class and not slot_is_online:
                        valid_slots.append(slot)
                
                if not valid_slots:
                    continue
                
                time_slot = random.choice(valid_slots)
                schedule['hari'] = day
                schedule['jam'] = f"{time_slot[0]} - {time_slot[1]}"
                
                # Untuk kelas online, tidak perlu cek ruangan fisik
                check_room = True
                if is_online_class:
                    room = 'Online'
                    check_room = False
                else:
                    department = schedule['kelas'][:2] if len(schedule['kelas']) >= 2 else 'default'
                    room = self.get_available_room(
                        department, 
                        day, 
                        time_slot[0], 
                        time_slot[1],
                        schedule.get('jumlah_mahasiswa', 0)
                    )
                
                if room:
                    schedule['ruangan'] = room
                    if not self.is_conflict(schedule, check_room_capacity=check_room):
                        success_count += 1
                        assigned = True
                        break
                else:
                    # Jika tidak ada ruangan, coba lagi
                    continue
            
            if not assigned:
                # Dapatkan alasan konflik
                conflict_reasons = self.get_conflict_reasons(schedule)
                failed_schedules.append({
                    'schedule': schedule,
                    'reasons': conflict_reasons
                })
                schedule['hari'] = ""
                schedule['jam'] = ""
                schedule['ruangan'] = ""
                failure_count += 1
        
        return success_count, failure_count, failed_schedules

    def validate_preferences(self):
        conflicts = []
        
        for lecturer, prefs in self.lecturer_preferences.items():
            available = prefs.get('available_days', [])
            online = prefs.get('online_days', [])
            
            for day in online:
                if day not in available:
                    conflicts.append(
                        f"{lecturer}: Hari {day} didefinisikan sebagai online tetapi tidak tersedia"
                    )
            
            for start, end in prefs.get('preferred_times_offline', []):
                if not self.is_valid_time_range(start, end):
                    conflicts.append(
                        f"{lecturer}: Waktu offline preferensi tidak valid ({start} - {end})"
                    )
            
            for start, end in prefs.get('preferred_times_online', []):
                if not self.is_valid_time_range(start, end):
                    conflicts.append(
                        f"{lecturer}: Waktu online preferensi tidak valid ({start} - {end})"
                    )
        
        return conflicts

    def get_lecturer_schedule(self, lecturer_name):
        return [s for s in self.fixed_schedules + self.generated_schedules 
                if s['dosen'] == lecturer_name]
    
    def randomize_all_rooms(self):
        all_schedules = self.fixed_schedules + self.generated_schedules
        all_schedules.sort(key=lambda x: x.get('jumlah_mahasiswa', 0), reverse=True)
    
        for sched in all_schedules:
            if sched.get('ruangan') == 'Online':
                continue
            
            if not sched.get('hari') or not sched.get('jam'):
                continue
            
            jam_parts = sched['jam'].split(' - ')
            if len(jam_parts) != 2:
                continue
            
            start, end = jam_parts
            student_count = sched.get('jumlah_mahasiswa', 0)
            department = sched['kelas'][:2] if isinstance(sched['kelas'], str) and len(sched['kelas']) >= 2 else 'default'
        
            room = self.get_available_room(
                department, 
                sched['hari'], 
                start, 
                end,
                student_count
            )
        
            if room:
                sched['ruangan'] = room


class ManualInputDialog(tk.Toplevel):
    def __init__(self, parent, generator, callback, schedule=None):
        super().__init__(parent)
        self.title("Edit Jadwal" if schedule else "Tambah Jadwal Manual")
        self.generator = generator
        self.callback = callback
        self.schedule = schedule
        
        # Form fields
        fields = [
            ("Hari:", 'hari_var', ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Online'], 0),
            ("Dosen:", 'dosen_var', None, 1),
            ("Mata Kuliah:", 'matkul_var', None, 2),
            ("Kelas:", 'kelas_var', None, 3),
            ("Ruangan:", 'ruangan_var', None, 4),
            ("Jam Mulai (HH:MM):", 'jam_mulai_var', None, 5),
            ("Jam Selesai:", 'jam_selesai_var', None, 6),
            ("SKS:", 'sks_var', None, 7),
            ("Semester:", 'semester_var', None, 8),
            ("Jumlah Mahasiswa:", 'mahasiswa_var', None, 9)
        ]
        
        self.vars = {}
        for i, (label, var_name, options, row) in enumerate(fields):
            ttk.Label(self, text=label).grid(row=row, column=0, padx=5, pady=5, sticky='e')
            
            var = tk.StringVar()
            self.vars[var_name] = var
            
            if options:
                ttk.OptionMenu(self, var, options[0], *options).grid(row=row, column=1, padx=5, pady=5, sticky='w')
            else:
                if var_name == 'dosen_var':
                    entry = ttk.Combobox(self, textvariable=var)
                    entry.grid(row=row, column=1, padx=5, pady=5, sticky='w')
                    self.dosen_combo = entry
                elif var_name == 'ruangan_var':
                    entry = ttk.Combobox(self, textvariable=var)
                    entry.grid(row=row, column=1, padx=5, pady=5, sticky='w')
                    self.ruangan_combo = entry
                elif var_name == 'jam_selesai_var':
                    label = ttk.Label(self, textvariable=var)
                    label.grid(row=row, column=1, padx=5, pady=5, sticky='w')
                else:
                    entry = ttk.Entry(self, textvariable=var)
                    entry.grid(row=row, column=1, padx=5, pady=5, sticky='w')
        
        # Tambahkan checkbox untuk jadwal tetap
        self.is_fixed_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(self, text="Jadwal Tetap (tidak diacak)", variable=self.is_fixed_var).grid(
            row=10, column=0, columnspan=2, sticky='w', padx=5, pady=5
        )
        
        ttk.Button(self, 
                  text="Simpan Perubahan" if schedule else "Tambah", 
                  command=self.save_schedule).grid(row=11, column=0, columnspan=2, pady=10)
        
        self.update_dropdowns()
        
        # Pasang trace untuk jam_mulai_var dan sks_var
        self.vars['jam_mulai_var'].trace_add("write", self.calculate_end_time)
        self.vars['sks_var'].trace_add("write", self.calculate_end_time)
        
        if schedule:
            # Set nilai untuk semua field secara langsung
            self.vars['hari_var'].set(schedule['hari'])
            self.vars['dosen_var'].set(schedule['dosen'])
            self.vars['matkul_var'].set(schedule['mata_kuliah'])
            self.vars['kelas_var'].set(schedule['kelas'])
            self.vars['ruangan_var'].set(schedule.get('ruangan', ''))
            self.vars['sks_var'].set(str(schedule['sks']))
            self.vars['semester_var'].set(str(schedule['semester']))
            self.vars['mahasiswa_var'].set(str(schedule.get('jumlah_mahasiswa', 0)))
            self.is_fixed_var.set(schedule.get('is_fixed', False))
            
            # Khusus jam: kita pecah
            jam_str = schedule['jam']
            if jam_str:
                parts = jam_str.split(' - ')
                if len(parts) == 2:
                    # Bersihkan dari (online)
                    start_clean = parts[0].replace('(online)', '').strip()
                    self.vars['jam_mulai_var'].set(start_clean)
        
        # Hitung jam selesai awal
        self.calculate_end_time()
        
    def calculate_end_time(self, *args):
        start_str = self.vars['jam_mulai_var'].get().strip()
        sks_str = self.vars['sks_var'].get().strip()
        
        if not start_str or not sks_str:
            self.vars['jam_selesai_var'].set('')
            return
        
        try:
            # Parse start time
            start_time = datetime.strptime(start_str, "%H:%M")
            sks = int(sks_str)
            
            # Calculate end time
            total_minutes = sks * 50
            end_time = start_time + timedelta(minutes=total_minutes)
            
            # Format end time
            end_time_str = end_time.strftime("%H:%M")
            self.vars['jam_selesai_var'].set(end_time_str)
        except Exception as e:
            # Jika parsing gagal, set kosong
            self.vars['jam_selesai_var'].set('')
    
    def update_dropdowns(self):
        # Update dosen combo
        self.dosen_combo['values'] = self.generator.lecturers
        
        # Update ruangan combo
        room_names = [room['nama'] for room in self.generator.available_rooms]
        room_names.append('Online')
        self.ruangan_combo['values'] = room_names
        
    def save_schedule(self):
        try:
            jam_mulai = self.vars['jam_mulai_var'].get().strip()
            jam_selesai = self.vars['jam_selesai_var'].get().strip()
            
            new_schedule = {
                'dosen': self.vars['dosen_var'].get(),
                'mata_kuliah': self.vars['matkul_var'].get(),
                'kelas': self.vars['kelas_var'].get(),
                'hari': self.vars['hari_var'].get(),
                'jam': f"{jam_mulai} - {jam_selesai}" if jam_mulai and jam_selesai else "",
                'semester': int(self.vars['semester_var'].get() or 0),
                'sks': int(self.vars['sks_var'].get() or 0),
                'ruangan': self.vars['ruangan_var'].get(),
                'jumlah_mahasiswa': int(self.vars['mahasiswa_var'].get() or 0),
                'is_fixed': self.is_fixed_var.get()  # Simpan status jadwal tetap
            }
            
            if new_schedule['hari'] == 'Online':
                new_schedule['ruangan'] = 'Online'
                if new_schedule['jam'] and not ("(online)" in new_schedule['jam'].lower()):
                    parts = new_schedule['jam'].split(' - ')
                    if len(parts) == 2:
                        new_schedule['jam'] = f"{parts[0]} (online) - {parts[1]} (online)"
            
            if not all([new_schedule['dosen'], new_schedule['mata_kuliah'], new_schedule['kelas']]):
                messagebox.showerror("Error", "Dosen, Mata Kuliah, dan Kelas harus diisi!")
                return
                
            if new_schedule['jam']:
                parts = new_schedule['jam'].split(' - ')
                if len(parts) != 2 or not self.generator.is_valid_time_range(parts[0], parts[1]):
                    messagebox.showerror("Error", "Format waktu tidak valid! Pastikan jam mulai valid dan SKS diisi.")
                    return
                
            # Untuk jadwal dari Excel, ubah menjadi manual setelah diedit
            if self.schedule and self.schedule.get('source') == 'excel':
                new_schedule['source'] = 'manual'
                new_schedule['excel_index'] = self.schedule.get('excel_index')
            
            if self.schedule:
                if self.generator.edit_schedule(self.schedule, new_schedule):
                    messagebox.showinfo("Sukses", "Jadwal berhasil diperbarui!")
                else:
                    messagebox.showerror("Error", "Gagal memperbarui jadwal")
            else:
                new_schedule['source'] = 'manual'
                self.generator.add_manual_schedule(new_schedule)
                messagebox.showinfo("Sukses", "Jadwal berhasil ditambahkan!")
            
            self.callback()
            self.destroy()
            
        except ValueError:
            messagebox.showerror("Error", "Pastikan SKS, Semester, dan Jumlah Mahasiswa berupa angka!")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menyimpan jadwal: {str(e)}")


class BreakTimeDialog(tk.Toplevel):
    def __init__(self, parent, generator, callback):
        super().__init__(parent)
        self.title("Tambah Waktu Istirahat Dosen")
        self.generator = generator
        self.callback = callback
        
        ttk.Label(self, text="Dosen:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.dosen_var = tk.StringVar()
        self.dosen_entry = ttk.Combobox(self, textvariable=self.dosen_var)
        self.dosen_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(self, text="Hari:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.hari_var = tk.StringVar()
        hari_options = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
        ttk.OptionMenu(self, self.hari_var, hari_options[0], *hari_options).grid(row=1, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(self, text="Waktu Mulai (HH:MM):").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        self.start_var = tk.StringVar(value="12:00")
        self.start_entry = ttk.Entry(self, textvariable=self.start_var)
        self.start_entry.grid(row=2, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(self, text="Waktu Selesai (HH:MM):").grid(row=3, column=0, padx=5, pady=5, sticky='e')
        self.end_var = tk.StringVar(value="13:00")
        self.end_entry = ttk.Entry(self, textvariable=self.end_var)
        self.end_entry.grid(row=3, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Button(self, text="Tambahkan", command=self.add_break).grid(row=4, column=0, columnspan=2, pady=10)
        
        self.dosen_entry['values'] = self.generator.lecturers
        
    def add_break(self):
        try:
            dosen = self.dosen_var.get()
            hari = self.hari_var.get()
            start = self.start_var.get()
            end = self.end_var.get()
            
            if not dosen:
                messagebox.showerror("Error", "Pilih dosen terlebih dahulu!")
                return
                
            if not re.match(r'^\d{1,2}:\d{2}$', start) or not re.match(r'^\d{1,2}:\d{2}$', end):
                messagebox.showerror("Error", "Format waktu tidak valid! Gunakan format HH:MM")
                return
                
            self.generator.add_lecturer_break(dosen, hari, start, end)
            messagebox.showinfo("Sukses", f"Waktu istirahat berhasil ditambahkan untuk {dosen} pada hari {hari} ({start} - {end})")
            self.destroy()
            self.callback()
            
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menambahkan waktu istirahat: {str(e)}")


class LecturerPreferenceDialog(tk.Toplevel):
    def __init__(self, parent, generator, callback):
        super().__init__(parent)
        self.title("Preferensi Jadwal Dosen")
        self.generator = generator
        self.callback = callback
        
        ttk.Label(self, text="Dosen:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.dosen_var = tk.StringVar()
        self.dosen_entry = ttk.Combobox(self, textvariable=self.dosen_var)
        self.dosen_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        self.dosen_entry['values'] = self.generator.lecturers
        self.dosen_entry.bind("<<ComboboxSelected>>", self.load_preference)
        
        # Changed to "Hari Tersedia"
        ttk.Label(self, text="Hari Tersedia:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.day_frame = ttk.Frame(self)
        self.day_frame.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        
        self.day_vars = {}
        days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
        for i, day in enumerate(days):
            var = tk.BooleanVar()
            cb = ttk.Checkbutton(self.day_frame, text=day, variable=var)
            cb.grid(row=0, column=i, padx=2)
            self.day_vars[day] = var
        
        ttk.Label(self, text="Hari Online:").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        self.online_day_frame = ttk.Frame(self)
        self.online_day_frame.grid(row=2, column=1, padx=5, pady=5, sticky='w')
        
        self.online_day_vars = {}
        for i, day in enumerate(days):
            var = tk.BooleanVar()
            cb = ttk.Checkbutton(self.online_day_frame, text=day, variable=var)
            cb.grid(row=0, column=i, padx=2)
            self.online_day_vars[day] = var
            
        ttk.Label(self, text="Preferensi Jam Offline (HH:MM):").grid(row=3, column=0, padx=5, pady=5, sticky='e')
        self.time_offline_frame = ttk.Frame(self)
        self.time_offline_frame.grid(row=3, column=1, padx=5, pady=5, sticky='w')
        
        self.time_offline_entries = []
        for i in range(2):
            start_var = tk.StringVar()
            end_var = tk.StringVar()
            ttk.Label(self.time_offline_frame, text=f"Slot {i+1}:").grid(row=i, column=0, padx=2)
            ttk.Entry(self.time_offline_frame, textvariable=start_var, width=8).grid(row=i, column=1, padx=2)
            ttk.Label(self.time_offline_frame, text="-").grid(row=i, column=2, padx=2)
            ttk.Entry(self.time_offline_frame, textvariable=end_var, width=8).grid(row=i, column=3, padx=2)
            self.time_offline_entries.append((start_var, end_var))
        
        ttk.Label(self, text="Preferensi Jam Online (HH:MM):").grid(row=4, column=0, padx=5, pady=5, sticky='e')
        self.time_online_frame = ttk.Frame(self)
        self.time_online_frame.grid(row=4, column=1, padx=5, pady=5, sticky='w')
        
        self.time_online_entries = []
        for i in range(2):
            start_var = tk.StringVar()
            end_var = tk.StringVar()
            ttk.Label(self.time_online_frame, text=f"Slot {i+1}:").grid(row=i, column=0, padx=2)
            ttk.Entry(self.time_online_frame, textvariable=start_var, width=8).grid(row=i, column=1, padx=2)
            ttk.Label(self.time_online_frame, text="-").grid(row=i, column=2, padx=2)
            ttk.Entry(self.time_online_frame, textvariable=end_var, width=8).grid(row=i, column=3, padx=2)
            self.time_online_entries.append((start_var, end_var))
        
        ttk.Label(self, text="Gunakan Break Time Tambahan:").grid(row=5, column=0, padx=5, pady=5, sticky='e')
        self.additional_breaks_var = tk.BooleanVar()
        ttk.Checkbutton(self, variable=self.additional_breaks_var).grid(row=5, column=1, sticky='w')
        
        ttk.Button(self, text="Simpan Preferensi", command=self.save_preference).grid(row=6, column=0, columnspan=2, pady=10)
        
    def load_preference(self, event=None):
        lecturer = self.dosen_var.get()
        if not lecturer:
            return
            
        pref = self.generator.lecturer_preferences.get(lecturer, {})
        
        # Load available days (hari tersedia)
        for day, var in self.day_vars.items():
            var.set(day in pref.get('available_days', []))
        
        for day, var in self.online_day_vars.items():
            var.set(day in pref.get('online_days', []))
        
        offline_times = pref.get('preferred_times_offline', [])
        for i, (start, end) in enumerate(offline_times):
            if i < len(self.time_offline_entries):
                self.time_offline_entries[i][0].set(start)
                self.time_offline_entries[i][1].set(end)
        
        online_times = pref.get('preferred_times_online', [])
        for i, (start, end) in enumerate(online_times):
            if i < len(self.time_online_entries):
                self.time_online_entries[i][0].set(start)
                self.time_online_entries[i][1].set(end)
        
        self.additional_breaks_var.set(pref.get('use_additional_breaks', False))
        
    def save_preference(self):
        try:
            lecturer = self.dosen_var.get()
            if not lecturer:
                messagebox.showerror("Error", "Pilih dosen terlebih dahulu!")
                return
                
            available_days = []
            for day, var in self.day_vars.items():
                if var.get():
                    available_days.append(day)
            
            online_days = []
            for day, var in self.online_day_vars.items():
                if var.get():
                    online_days.append(day)
            
            preferred_times_offline = []
            for start_var, end_var in self.time_offline_entries:
                start = start_var.get().strip()
                end = end_var.get().strip()
                if start and end:
                    if not re.match(r'^\d{1,2}:\d{2}$', start) or not re.match(r'^\d{1,2}:\d{2}$', end):
                        messagebox.showerror("Error", "Format waktu offline tidak valid! Gunakan format HH:MM")
                        return
                    preferred_times_offline.append((start, end))
            
            preferred_times_online = []
            for start_var, end_var in self.time_online_entries:
                start = start_var.get().strip()
                end = end_var.get().strip()
                if start and end:
                    if not re.match(r'^\d{1,2}:\d{2}$', start) or not re.match(r'^\d{1,2}:\d{2}$', end):
                        messagebox.showerror("Error", "Format waktu online tidak valid! Gunakan format HH:MM")
                        return
                    preferred_times_online.append((start, end))
            
            additional_breaks = self.additional_breaks_var.get()
            
            self.generator.add_lecturer_preference(
                lecturer, 
                available_days, 
                preferred_times_offline, 
                preferred_times_online, 
                online_days, 
                additional_breaks
            )
            messagebox.showinfo("极kses", f"Preferensi berhasil disimpan untuk {lecturer}")
            self.destroy()
            self.callback()
            
        except Exception as e:
            messagebox.showerror("Error", f"极gagal menyimpan preferensi: {str(e)}")


class ScheduleApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Nusaputra Schedule Generator")
        self.root.geometry("1200x800")
        self.root.minsize(1000, 700)  # Set minimum window size
        self.generator = ScheduleGenerator()
        
        # Load style
        self.setup_styles()
        
        if not self.generator.load_cache():
            try:
                if os.path.exists("data/rooms.json"):
                    self.generator.load_rooms("data/rooms.json")
            except:
                pass
        
        self.setup_auto_save()
        atexit.register(self.generator.save_cache)
        
        self.sort_order_hari = 'asc'
        self.current_filter_hari = None
        self.current_filter_mode = None
        self.selected_schedule = None
        self.create_widgets()
        self.load_ui_state()

    def setup_styles(self):
        """Configure visual styles for consistent UI"""
        style = ttk.Style()
        style.configure("TFrame", background="#f0f0f0")
        style.configure("TLabel", background="#f0f0f0")
        style.configure("TButton", padding=5)
        style.configure("Title.TLabel", font=("Arial", 16, "bold"), foreground="#2c3e50")
        style.configure("Section.TLabel", font=("Arial", 10, "bold"), foreground="#3498db")
        style.configure("Status.TLabel", font=("Arial", 9), foreground="#7f8c8d")
        style.configure("Treeview", rowheight=25)
        style.configure("Treeview.Heading", font=("Arial", 9, "bold"))
        style.map("Treeview", background=[("selected", "#3498db")])
        style.configure("Action.TButton", padding=5, font=("Arial", 9))

    def setup_auto_save(self):
        self.generator.save_cache()
        self.root.after(300000, self.setup_auto_save)

    def save_ui_state(self):
        state = {
            'window_geometry': self.root.geometry(),
            'selected_lecturer': self.lecturer_var.get(),
            'hari_filter': self.hari_var.get(),
            'mode_filter': self.mode_var.get(),
            'sort_order': self.sort_order_hari
        }
        self.generator.save_ui_state(state)

    def load_ui_state(self):
        state = self.generator.load_ui_state()
        if state:
            try:
                self.root.geometry(state.get('window_geometry', '1200x800'))
                self.lecturer_var.set(state.get('selected_lecturer', ''))
                self.hari_var.set(state.get('hari_filter', 'Semua'))
                self.mode_var.set(state.get('mode_filter', 'Semua'))
                self.sort_order_hari = state.get('sort_order', 'asc')
                
                if self.sort_order_hari == 'desc':
                    self.sort_hari_btn.config(text="Sort Hari (Z-A)")
                
                if self.lecturer_var.get():
                    self.show_lecturer_schedule()
            except:
                pass

    def create_widgets(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # Header section
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(header_frame, text="Penjadwalan Kuliah Universitas Nusaputra", 
                 style="Title.TLabel").pack()
        
        # Main content area with two columns
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Left column - Filters and controls dengan scrollbar
        left_container = ttk.Frame(content_frame)
        left_container.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        
        # Buat canvas dan scrollbar vertikal
        canvas = tk.Canvas(left_container)
        scrollbar = ttk.Scrollbar(left_container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Tambahkan binding untuk mouse wheel
        canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        canvas.bind("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))  # Untuk Linux
        canvas.bind("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))   # Untuk Linux
        scrollable_frame.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        scrollable_frame.bind("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
        scrollable_frame.bind("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))
        
        left_frame = ttk.LabelFrame(scrollable_frame, text="Kontrol & Filter", padding=20)
        left_frame.pack(fill=tk.X, pady=5)
        
        # ========== LEFT COLUMN CONTROLS ==========
        
        # Data Management Section
        data_frame = ttk.LabelFrame(left_frame, text="Manajemen Data", padding=5)
        data_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(data_frame, text="📂 Load Data Excel", command=self.load_excel_data,
                  style="Action.TButton").pack(fill=tk.X, pady=2)
        ttk.Button(data_frame, text="🏢 Load Data Ruangan", command=self.load_room_data,
                  style="Action.TButton").pack(fill=tk.X, pady=2)
        ttk.Button(data_frame, text="💾 Simpan Semua Jadwal", command=self.save_schedule_all,
                  style="Action.TButton").pack(fill=tk.X, pady=2)
        
        # Filter Section
        filter_frame = ttk.LabelFrame(left_frame, text="Filter Jadwal", padding=5)
        filter_frame.pack(fill=tk.X, pady=5)
        
        # Dosen Selection
        ttk.Label(filter_frame, text="Pilih Dosen:").pack(anchor=tk.W, pady=2)
        self.lecturer_var = tk.StringVar()
        self.lecturer_dropdown = ttk.Combobox(filter_frame, 
                                             textvariable=self.lecturer_var, 
                                             width=25)
        self.lecturer_dropdown.pack(fill=tk.X, pady=2)
        self.lecturer_dropdown.bind("<<ComboboxSelected>>", self.show_lecturer_schedule)
        
        # Filter Hari
        ttk.Label(filter_frame, text="Filter Hari:").pack(anchor=tk.W, pady=2)
        self.hari_var = tk.StringVar()
        hari_options = ['Semua', 'Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Online']
        self.hari_dropdown = ttk.Combobox(filter_frame, 
                                        textvariable=self.hari_var,
                                        values=hari_options,
                                        state='readonly',
                                        width=12)
        self.hari_dropdown.pack(fill=tk.X, pady=2)
        self.hari_dropdown.set('Semua')
        self.hari_dropdown.bind("<<ComboboxSelected>>", self.apply_filters)
        
        # Filter Mode
        ttk.Label(filter_frame, text="Filter Mode:").pack(anchor=tk.W, pady=2)
        self.mode_var = tk.StringVar()
        mode_options = ['Semua', 'Online', 'Offline']
        self.mode_dropdown = ttk.Combobox(filter_frame, 
                                        textvariable=self.mode_var,
                                        values=mode_options,
                                        state='readonly',
                                        width=12)
        self.mode_dropdown.pack(fill=tk.X, pady=2)
        self.mode_dropdown.set('Semua')
        self.mode_dropdown.bind("<<ComboboxSelected>>", self.apply_filters)
        
        # Sort Button
        self.sort_hari_btn = ttk.Button(filter_frame, 
                                      text="🔽 Sort Hari (A-Z)", 
                                      command=self.toggle_sort_hari)
        self.sort_hari_btn.pack(fill=tk.X, pady=5)
        
        # Schedule Actions Section
        action_frame = ttk.LabelFrame(left_frame, text="Aksi Jadwal", padding=5)
        action_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(action_frame, text="🎲 Acak Jadwal Baru", command=self.randomize_schedule,
                  style="Action.TButton").pack(fill=tk.X, pady=2)
        ttk.Button(action_frame, text="🔄 Acak Ulang Semua", command=lambda: self.randomize_schedule(True),
                  style="Action.TButton").pack(fill=tk.X, pady=2)
        ttk.Button(action_frame, text="✏️ Tambah Jadwal Manual", command=self.show_manual_input,
                  style="Action.TButton").pack(fill=tk.X, pady=2)
        ttk.Button(action_frame, text="🔄 Acak Ruangan", command=self.generate_rooms,
                  style="Action.TButton").pack(fill=tk.X, pady=2)
        
        # Conflict Management Section
        conflict_frame = ttk.LabelFrame(left_frame, text="Manajemen Konflik", padding=5)
        conflict_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(conflict_frame, text="⚠️ Cek Konflik", command=self.show_conflicts,
                  style="Action.TButton").pack(fill=tk.X, pady=2)
        ttk.Button(conflict_frame, text="🔧 Atasi Konflik", command=self.resolve_conflicts,
                  style="Action.TButton").pack(fill=tk.X, pady=2)
        
        # Lecturer Preferences Section
        pref_frame = ttk.LabelFrame(left_frame, text="Preferensi Dosen", padding=5)
        pref_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(pref_frame, text="👨‍🏫 Kelola Preferensi", command=self.show_lecturer_preference,
                  style="Action.TButton").pack(fill=tk.X, pady=2)
        ttk.Button(pref_frame, text="✅ Validasi Preferensi", command=self.validate_preferences,
                  style="Action.TButton").pack(fill=tk.X, pady=2)
        ttk.Button(pref_frame, text="☕ Waktu Istirahat", command=self.show_break_time_dialog,
                  style="Action.TButton").pack(fill=tk.X, pady=2)
        
        # ========== RIGHT COLUMN SCHEDULE DISPLAY ==========
        
        # Right column - Schedule display
        right_frame = ttk.Frame(content_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Data section
        data_display_frame = ttk.LabelFrame(right_frame, text="Data Jadwal")
        data_display_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Treeview with scrollbars
        tree_container = ttk.Frame(data_display_frame)
        tree_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        columns = [
            ('Sumber', 80), ('Hari', 80), ('Dosen', 150), ('Mata Kuliah', 180), ('Kelas', 60), 
            ('Ruangan', 80), ('Jam', 120), ('SKS', 40), ('Semester', 70), ('Mahasiswa', 70)
        ]
        
        self.schedule_tree = ttk.Treeview(tree_container, columns=[c[0] for c in columns], show='headings', height=15)
        
        for col, width in columns:
            self.schedule_tree.heading(col, text=col)
            self.schedule_tree.column(col, width=width, anchor='center')
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.schedule_tree.yview)
        hsb = ttk.Scrollbar(tree_container, orient="horizontal", command=self.schedule_tree.xview)
        self.schedule_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Layout
        self.schedule_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)
        
        self.schedule_tree.bind('<<TreeviewSelect>>', self.on_schedule_select)

        # Action buttons for selected schedule
        action_btn_frame = ttk.Frame(data_display_frame)
        action_btn_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(action_btn_frame, text="✏️ Edit Jadwal Terpilih", 
                  command=self.edit_selected_schedule).pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)
        ttk.Button(action_btn_frame, text="🗑️ Hapus Jadwal Terpilih", 
                  command=self.delete_selected_schedule).pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)
        ttk.Button(action_btn_frame, text="📌 Tandai sebagai Tetap", 
                  command=self.toggle_fixed_schedule).pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)
        
        # NEW: Save buttons frame
        save_btn_frame = ttk.Frame(data_display_frame)
        save_btn_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(save_btn_frame, text="💾 Simpan Jadwal Dosen Ini", 
                  command=self.save_current_lecturer_schedule).pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)
        ttk.Button(save_btn_frame, text="💾 Simpan Semua Jadwal", 
                  command=self.save_schedule_all).pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)
        
        # Status Bar
        status_frame = ttk.Frame(right_frame)
        status_frame.pack(fill=tk.X, pady=(5, 0))
        
        self.status_var = tk.StringVar()
        self.status_var.set("Siap | Pilih dosen untuk melihat jadwal")
        ttk.Label(status_frame, textvariable=self.status_var, style="Status.TLabel",
                 anchor=tk.W).pack(fill=tk.X)
        
        # Context menu
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="✏️ Edit Jadwal", command=self.edit_selected_schedule)
        self.context_menu.add_command(label="🗑️ Hapus Jadwal", command=self.delete_selected_schedule)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="📌 Tandai sebagai Tetap", command=self.toggle_fixed_schedule)
        self.schedule_tree.bind("<Button-3>", self.show_context_menu)
        
        if self.generator.lecturers:
            self.lecturer_dropdown['values'] = self.generator.lecturers
            if not self.lecturer_var.get() and self.generator.lecturers:
                self.lecturer_var.set(self.generator.lecturers[0])
                self.show_lecturer_schedule()

    def save_current_lecturer_schedule(self):
        lecturer = self.lecturer_var.get()
        if not lecturer:
            messagebox.showwarning("Peringatan", "Pilih dosen terlebih dahulu!")
            return
            
        all_schedules = self.generator.fixed_schedules + self.generator.generated_schedules
        lecturer_schedules = [s for s in all_schedules if s['dosen'] == lecturer]
        
        if not lecturer_schedules:
            messagebox.showinfo("Info", f"Tidak ada jadwal untuk dosen {lecturer}")
            return
            
        folder = filedialog.askdirectory(title="Pilih Folder Output")
        if folder:
            template_path = "templates/schedule_template.xlsx"
            if not os.path.exists(template_path):
                template_path = filedialog.askopenfilename(title="Pilih Template Excel", filetypes=[("Excel Files", "*.xlsx")])
            
            if template_path:
                # Create safe filename
                safe_lecturer_name = re.sub(r'[^a-zA-Z0-9]', '_', lecturer)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_filename = f"Jadwal_{safe_lecturer_name}_{timestamp}.xlsx"
                output_path = os.path.join(folder, output_filename)
                
                try:
                    wb = load_workbook(template_path)
                    sheet = wb.active
                    
                    headers = ["Hari", "Mata Kuliah", "Kelas", "Ruangan", "Jam", "SKS", "Semester", "Dosen", "Jumlah Mahasiswa"]
                    for col, header in enumerate(headers, start=1):
                        sheet.cell(row=3, column=col, value=header)
                    
                    for row_idx, sched in enumerate(lecturer_schedules, start=4):
                        sheet.cell(row=row_idx, column=1, value=sched['hari'])
                        sheet.cell(row=row_idx, column=2, value=sched['mata_kuliah'])
                        sheet.cell(row=row_idx, column=3, value=sched['kelas'])
                        sheet.cell(row=row_idx, column=4, value=sched.get('ruangan', ''))
                        sheet.cell(row=row_idx, column=5, value=sched['jam'])
                        sheet.cell(row=row_idx, column=6, value=sched['sks'])
                        sheet.cell(row=row_idx, column=7, value=sched['semester'])
                        sheet.cell(row=row_idx, column=8, value=sched['dosen'])
                        sheet.cell(row=row_idx, column=9, value=sched.get('jumlah_mahasiswa', ''))
                        
                    wb.save(output_path)
                    messagebox.showinfo("Sukses", f"Jadwal untuk dosen {lecturer} berhasil disimpan di:\n{output_path}")
                    # Open the output folder
                    os.startfile(folder)
                except Exception as e:
                    messagebox.showerror("Error", f"Gagal menyimpan: {str(e)}")   

    def load_data_wrapper(self):
        """Wrapper untuk memilih tipe data yang akan dimuat"""
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="Load Data Excel", command=self.load_excel_data)
        menu.add_command(label="Load Data Ruangan", command=self.load_room_data)
        menu.post(self.root.winfo_pointerx(), self.root.winfo_pointery())

    def toggle_fixed_schedule(self):
        """Toggle status jadwal tetap"""
        if self.selected_schedule:
            self.selected_schedule['is_fixed'] = not self.selected_schedule.get('is_fixed', False)
            status = "DITETAPKAN" if self.selected_schedule['is_fixed'] else "TIDAK TETAP"
            self.status_var.set(f"Status jadwal diubah: {status}")
            self.show_lecturer_schedule()

    def show_context_menu(self, event):
        item = self.schedule_tree.identify_row(event.y)
        if item:
            self.schedule_tree.selection_set(item)
            self.on_schedule_select(None)  # Untuk memilih jadwal
            self.context_menu.post(event.x_root, event.y_root)

    def load_excel_data(self):
        path = filedialog.askopenfilename(title="Pilih File Excel", filetypes=[("Excel Files", "*.xlsx")])
        if path and self.generator.load_data(path):
            self.lecturer_dropdown["values"] = self.generator.lecturers
            if self.generator.lecturers:
                self.lecturer_var.set(self.generator.lecturers[0])
                self.show_lecturer_schedule()
            self.status_var.set("Data jadwal berhasil dimuat dari Excel")
            self.save_ui_state()

    def load_room_data(self):
        path = filedialog.askopenfilename(title="Pilih File Ruangan (JSON)", filetypes=[("JSON Files", "*.json")])
        if path and self.generator.load_rooms(path):
            self.status_var.set("Data ruangan berhasil dimuat")
            self.save_ui_state()

    def show_lecturer_schedule(self, event=None):
        lecturer = self.lecturer_var.get()
        self.schedule_tree.delete(*self.schedule_tree.get_children())
        
        if not lecturer:
            self.status_var.set("Pilih dosen terlebih dahulu")
            return
            
        all_schedules = self.generator.fixed_schedules + self.generator.generated_schedules
        filtered_schedules = [s for s in all_schedules if s['dosen'] == lecturer]
        
        hari_filter = self.hari_var.get()
        if hari_filter and hari_filter != 'Semua':
            if hari_filter == 'Online':
                filtered_schedules = [s for s in filtered_schedules if s.get('ruangan') == 'Online']
            else:
                filtered_schedules = [s for s in filtered_schedules if s['hari'] == hari_filter]
        
        mode_filter = self.mode_var.get()
        if mode_filter == 'Online':
            filtered_schedules = [s for s in filtered_schedules if s.get('ruangan') == 'Online']
        elif mode_filter == 'Offline':
            filtered_schedules = [s for s in filtered_schedules if s.get('ruangan') != 'Online']
        
        if self.sort_order_hari == 'asc':
            filtered_schedules.sort(key=lambda x: x['hari'])
        else:
            filtered_schedules.sort(key=lambda x: x['hari'], reverse=True)
        
        for s in filtered_schedules:
            # Tentukan sumber jadwal
            source = "Excel" if s.get('source') == 'excel' else "Manual"
            
            # Tandai jadwal tetap dengan ikon khusus
            tags = ("fixed",) if s.get('is_fixed', False) else ()
            
            self.schedule_tree.insert('', 'end', values=(
                source,
                s['hari'],
                s['dosen'],
                s['mata_kuliah'],
                s['kelas'],
                s.get('ruangan', ''),
                s['jam'],
                s['sks'],
                s['semester'],
                s.get('jumlah_mahasiswa', '')
            ), tags=tags)
        
        # Konfigurasi tag untuk jadwal tetap
        self.schedule_tree.tag_configure("fixed", background="#e0f7fa")
        
        status_text = f"Menampilkan {len(filtered_schedules)} jadwal untuk {lecturer}"
        if any(s.get('is_fixed', False) for s in filtered_schedules):
            status_text += " | Jadwal biru: Jadwal tetap"
        self.status_var.set(status_text)
        
        self.save_ui_state()

    def show_failed_schedules_dialog(self, failed_schedules):
        dialog = tk.Toplevel(self.root)
        dialog.title("Jadwal yang Gagal Diacak")
        dialog.geometry("800x500")
        dialog.transient(self.root)
        dialog.grab_set()
        
        main_frame = ttk.Frame(dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        ttk.Label(main_frame, text="Jadwal berikut gagal diacak karena konflik:", 
                 font=("Arial", 10, "bold")).pack(pady=(0, 10))
        
        # Frame untuk treeview
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = [('Dosen', 150), ('Mata Kuliah', 200), ('Kelas', 80), ('Konflik', 300)]
        tree = ttk.Treeview(tree_frame, columns=[c[0] for c in columns], show='headings', height=10)
        
        for col, width in columns:
            tree.heading(col, text=col)
            tree.column(col, width=width, anchor='w')
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Isi data
        for item in failed_schedules:
            schedule = item['schedule']
            reasons = "\n".join(item['reasons']) if item['reasons'] else "Tidak diketahui"
            tree.insert('', 'end', values=(
                schedule['dosen'],
                schedule['mata_kuliah'],
                schedule['kelas'],
                reasons
            ))
        
        # Button frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(btn_frame, text="Tutup", command=dialog.destroy).pack(side=tk.RIGHT)

    def apply_filters(self, event=None):
        self.show_lecturer_schedule()

    def toggle_sort_hari(self):
        if self.sort_order_hari == 'asc':
            self.sort_order_hari = 'desc'
            self.sort_hari_btn.config(text="Sort Hari (Z-A)")
        else:
            self.sort_order_hari = 'asc'
            self.sort_hari_btn.config(text="Sort Hari (A-Z)")
        self.show_lecturer_schedule()
        self.save_ui_state()

    def generate_rooms(self):
        self.generator.randomize_all_rooms()
        self.show_lecturer_schedule()
        self.status_var.set("Ruangan berhasil diacak ulang untuk semua jadwal!")
        self.save_ui_state()

    def save_schedule_all(self):
        all_sched = self.generator.fixed_schedules + self.generator.generated_schedules
        folder = filedialog.askdirectory(title="Pilih Folder Output")
        if folder:
            template_path = "templates/schedule_template.xlsx"
            if not os.path.exists(template_path):
                template_path = filedialog.askopenfilename(title="Pilih Template Excel", filetypes=[("Excel Files", "*.xlsx")])
            
            if template_path:
                out = self.generator.save_to_excel(all_sched, template_path, folder)
                if out:
                    messagebox.showinfo("Sukses", f"Jadwal disimpan di:\n{out}")
                    os.startfile(folder)
                    self.save_ui_state()

    def show_conflicts(self):
        conflicts = self.generator.find_all_conflicts()
        conflict_window = tk.Toplevel(self.root)
        conflict_window.title("Konflik Jadwal")
        conflict_window.geometry("1000x600")
        
        notebook = ttk.Notebook(conflict_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        conflict_types = {
            'Dosen': ['lecturer'],
            'Ruangan': ['room', 'capacity'],
            'Kelas': ['class'],
            'Preferensi': ['break_time', 'online_day', 'preference']
        }
        
        for title, types in conflict_types.items():
            frame = ttk.Frame(notebook)
            notebook.add(frame, text=title)
            
            tree = ttk.Treeview(frame, 
                               columns=('Tipe', 'Entitas', 'Hari', 'Waktu', 'Detail1', 'Detail2', 'Solusi'), 
                               show='headings',
                               height=15)
            
            scroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            scroll.pack(side=tk.RIGHT, fill=tk.Y)
            tree.configure(yscrollcommand=scroll.set)
            tree.pack(fill=tk.BOTH, expand=True)
            
            tree.heading('Tipe', text='Tipe')
            tree.heading('Entitas', text='Entitas')
            tree.heading('Hari', text='Hari')
            tree.heading('Waktu', text='Waktu')
            tree.heading('Detail1', text='Detail 1')
            tree.heading('Detail2', text='Detail 2')
            tree.heading('Solusi', text='Solusi')
            
            tree.column('Tipe', width=100, anchor='w')
            tree.column('Entitas', width=100, anchor='w')
            tree.column('Hari', width=80, anchor='center')
            tree.column('Waktu', width=100, anchor='center')
            tree.column('Detail1', width=180, anchor='w')
            tree.column('Detail2', width=180, anchor='w')
            tree.column('Solusi', width=200, anchor='w')
            
            for c_type in types:
                for conflict in conflicts.get(c_type, []):
                    solutions = self.generator.suggest_conflict_resolutions(conflict)
                    solution_text = solutions[0] if solutions else "Perlu penyesuaian manual"
                    
                    if c_type == 'lecturer':
                        tree.insert('', 'end', values=(
                            conflict['conflict_type'],
                            conflict['dosen'],
                            conflict['hari'],
                            conflict['waktu'],
                            f"{conflict['schedule1']['mata_kuliah']} ({conflict['schedule1']['kelas']})",
                            f"{conflict['schedule2']['mata_kuliah']} ({conflict['schedule2']['kelas']})",
                            solution_text
                        ))
                    elif c_type == 'room':
                        tree.insert('', 'end', values=(
                            conflict['conflict_type'],
                            conflict['ruangan'],
                            conflict['hari'],
                            conflict['waktu'],
                            f"{conflict['schedule1']['mata_kuliah']} ({conflict['schedule1']['kelas']})",
                            f"{conflict['schedule2']['mata_kuliah']} ({conflict['schedule2']['kelas']})",
                            solution_text
                        ))
                    elif c_type == 'capacity':
                        tree.insert('', 'end', values=(
                            conflict['conflict_type'],
                            f"{conflict['ruangan']} (Kap: {conflict['kapasitas']})",
                            conflict['schedule']['hari'],
                            conflict['schedule']['jam'],
                            f"{conflict['schedule']['mata_kuliah']} ({conflict['schedule']['kelas']})",
                            f"Mahasiswa: {conflict['mahasiswa']}",
                            solution_text
                        ))
                    elif c_type == 'break_time':
                        tree.insert('', 'end', values=(
                            conflict['conflict_type'],
                            conflict['dosen'],
                            conflict['hari'],
                            conflict['waktu'],
                            f"{conflict['schedule']['mata_kuliah']} ({conflict['schedule']['kelas']})",
                            "Waktu istirahat",
                            solution_text
                        ))
                    elif c_type == 'online_day':
                        tree.insert('', 'end', values=(
                            conflict['conflict_type'],
                            conflict['dosen'],
                            conflict['hari'],
                            conflict['waktu'],
                            f"{conflict['schedule']['mata_kuliah']} ({conflict['schedule']['kelas']})",
                            f"Ruangan: {conflict['ruangan']}",
                            solution_text
                        ))
                    elif c_type == 'preference':
                        tree.insert('', 'end', values=(
                            conflict['conflict_type'],
                            conflict['dosen'],
                            conflict['hari'],
                            conflict['waktu'],
                            f"{conflict['schedule']['mata_kuliah']} ({conflict['schedule']['kelas']})",
                            "",
                            solution_text
                        ))

    def show_manual_input(self):
        ManualInputDialog(self.root, self.generator, self.show_lecturer_schedule)

    def on_schedule_select(self, event):
        selected = self.schedule_tree.selection()
        if selected:
            item = self.schedule_tree.item(selected[0])
            values = item['values']
            
            # Kolom: 
            # 0: Sumber, 1: Hari, 2: Dosen, 3: Mata Kuliah, 4: Kelas, 
            # 5: Ruangan, 6: Jam, 7: SKS, 8: Semester, 9: Mahasiswa
            
            lecturer = values[2]
            hari = values[1]
            matkul = values[3]
            kelas = values[4]
            ruangan = values[5]
            jam = values[6]
            
            all_schedules = self.generator.fixed_schedules + self.generator.generated_schedules
            
            for s in all_schedules:
                if (s['dosen'] == lecturer and 
                    s['hari'] == hari and
                    s['mata_kuliah'] == matkul and
                    s['kelas'] == kelas and
                    s.get('ruangan', '') == ruangan and
                    s['jam'] == jam):
                    self.selected_schedule = s
                    return
                    
            self.selected_schedule = None
        else:
            self.selected_schedule = None

    def edit_selected_schedule(self):
        if not self.selected_schedule:
            messagebox.showwarning("Peringatan", "Pilih jadwal yang akan diedit!")
            return
            
        ManualInputDialog(self.root, self.generator, self.show_lecturer_schedule, self.selected_schedule)

    def resolve_conflicts(self):
        resolved = self.generator.auto_resolve_conflicts()
        if resolved > 0:
            messagebox.showinfo("Sukses", f"Berhasil menyelesaikan {resolved} konflik!")
            self.show_lecturer_schedule()
            self.save_ui_state()
        else:
            messagebox.showinfo("Info", "Tidak ada konflik yang bisa diselesaikan secara otomatis")
            
    def randomize_schedule(self, reshuffle_existing=False):
        if not self.generator.lecturers:
            messagebox.showwarning("Peringatan", "Load data dosen terlebih dahulu!")
            return
        
        if reshuffle_existing:
            if not messagebox.askyesno("Konfirmasi", 
                                      "Apakah Anda yakin ingin mengacak ulang SEMUA jadwal?\n"
                                      "Ini akan menghapus semua jadwal yang sudah dibuat sebelumnya."):
                return
        
        success_count, failure_count, failed_schedules = self.generator.randomize_schedule(reshuffle_existing)
    
        if success_count == 0 and failure_count == 0:
            message = "Semua jadwal sudah memiliki waktu. Tidak ada yang diacak."
        elif failure_count > 0:
            self.show_failed_schedules_dialog(failed_schedules)
            message = (f"Berhasil mengacak {success_count} jadwal!\n" 
                    f"{failure_count} jadwal gagal diacak (lihat detail konflik).")
        else:
            message = f"Berhasil mengacak {success_count} jadwal!"
    
        messagebox.showinfo("Hasil Pengacakan", message)
        self.show_lecturer_schedule()
        self.save_ui_state()
            
    def show_lecturer_preference(self):
        LecturerPreferenceDialog(self.root, self.generator, self.show_lecturer_schedule)

    def validate_preferences(self):
        conflicts = self.generator.validate_preferences()
        if conflicts:
            msg = "Konflik preferensi ditemukan:\n\n" + "\n".join(conflicts)
            messagebox.showwarning("Validasi Preferensi", msg)
        else:
            messagebox.showinfo("Validasi Preferensi", "Semua preferensi dosen valid!")

    def show_break_time_dialog(self):
        BreakTimeDialog(self.root, self.generator, self.show_lecturer_schedule)

    def delete_selected_schedule(self):
        if not self.selected_schedule:
            messagebox.showwarning("Peringatan", "Pilih jadwal yang akan dihapus!")
            return
            
        if self.generator.remove_schedule(self.selected_schedule):
            self.status_var.set("Jadwal berhasil dihapus")
            self.show_lecturer_schedule()
            self.save_ui_state()
        else:
            messagebox.showerror("Error", "Gagal menghapus jadwal")

    def on_closing(self):
        self.save_ui_state()
        self.generator.save_cache()
        self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    try:
        app = ScheduleApp(root)
        root.protocol("WM_DELETE_WINDOW", app.on_closing)
        root.mainloop()
    except Exception as e:
        print(f"Application crashed: {e}")
        traceback.print_exc()
        app.generator.save_cache()
        messagebox.showerror("Error", f"Aplikasi mengalami error: {e}\nData telah disimpan di cache.")